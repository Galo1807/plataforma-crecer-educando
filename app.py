from flask import Flask, render_template, request, redirect, url_for, session, flash
from werkzeug.utils import secure_filename
import os
import sqlite3
from flask import send_file
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import pandas as pd
from openpyxl import Workbook
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

app = Flask(__name__)
app.secret_key = 'editorial_secret_key'
UPLOAD_FOLDER = 'static/images'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

DATABASE = 'database/editorial.db'

def precio_del_libro(nombre):
    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    c.execute("SELECT precio FROM libros WHERE nombre = ?", (nombre,))
    resultado = c.fetchone()
    conn.close()
    return resultado[0] if resultado else 0.0

def get_db_connection():
    conn = sqlite3.connect('database/editorial.db')
    conn.row_factory = sqlite3.Row
    return conn

# Obtener cliente
def obtener_cliente(cliente_id):
    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    c.execute("SELECT * FROM clientes WHERE id = ?", (cliente_id,))
    cliente = c.fetchone()
    conn.close()
    return cliente

# Obtener entregas
def obtener_entregas(cliente_id):
    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    c.execute("""
        SELECT id, libro, cantidad, precio_unitario, fecha, nota
        FROM ventas
        WHERE cliente_id = ?
    """, (cliente_id,))
    entregas = c.fetchall()
    conn.close()
    return entregas

def guardar_entrega(cliente_id, fecha, nota, materias, cantidades, descuentos, precios, imagen):
    conn = sqlite3.connect('database/editorial.db')
    cursor = conn.cursor()

    nombre_imagen = None
    if imagen and imagen.filename:
        nombre_imagen = secure_filename(imagen.filename)
        ruta_directorio = os.path.join('static', 'notas')
        os.makedirs(ruta_directorio, exist_ok=True)
        imagen.save(os.path.join(ruta_directorio, nombre_imagen))

    for materia, cantidad, descuento, precio in zip(materias, cantidades, descuentos, precios):
        if int(cantidad) > 0:
            cursor.execute("""
                INSERT INTO ventas (cliente_id, fecha, nota, libro, cantidad, precio_unitario, descuento, imagen)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (cliente_id, fecha, nota, materia, int(cantidad), float(precio), int(descuento), nombre_imagen))

    conn.commit()
    conn.close()


# Obtener devoluciones
def obtener_devoluciones(cliente_id):
    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    c.execute("SELECT * FROM devoluciones WHERE cliente_id = ?", (cliente_id,))
    devoluciones = c.fetchall()
    conn.close()
    return devoluciones

# Obtener abonos
def obtener_abonos(cliente_id):
    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    c.execute("SELECT * FROM abonos WHERE cliente_id = ?", (cliente_id,))
    abonos = c.fetchall()
    conn.close()
    return abonos

# Guardar devolución
def guardar_devolucion(cliente_id, fecha, materia, cantidad):
    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    c.execute("""
        INSERT INTO devoluciones (cliente_id, libro, cantidad, fecha)
        VALUES (?, ?, ?, ?)
    """, (cliente_id, materia, cantidad, fecha))
    conn.commit()
    conn.close()

# Guardar abono
def guardar_abono(cliente_id, fecha, monto, comentario, comprobante=None):
    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    c.execute("""
        INSERT INTO abonos (cliente_id, monto, fecha, comentario, comprobante)
        VALUES (?, ?, ?, ?, ?)
    """, (cliente_id, monto, fecha, comentario, comprobante))
    conn.commit()
    conn.close()

# Generar PDF de estado de cuenta
def generar_estado_cuenta_pdf(cliente, entregas, devoluciones, abonos):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, height - 50, f"Estado de Cuenta - Cliente: {cliente[1]}")
    p.setFont("Helvetica", 12)

    y = height - 80
    p.drawString(50, y, f"ID Cliente: {cliente[0]}")
    y -= 30

    p.setFont("Helvetica-Bold", 12)
    p.drawString(50, y, "Entregas:")
    y -= 20
    total_entregado = 0
    for e in entregas:
        libro = e[1]
        cantidad = int(e[2])
        precio = float(e[3])
        fecha = e[4]
        nota = e[5] if e[5] else "Sin Nota"
        subtotal = cantidad * precio
        total_entregado += subtotal

        linea = f"Nota {nota} - {fecha} - {libro}: {cantidad} x ${precio:.2f} = ${subtotal:.2f}"
        p.setFont("Helvetica", 10)
        p.drawString(60, y, linea)
        y -= 15


    y -= 15
    p.setFont("Helvetica-Bold", 10)
    p.drawString(50, y, "Devoluciones:")
    y -= 20
    total_devoluciones = 0
    for d in devoluciones:
        total_devoluciones += d[3]
        p.drawString(60, y, f"{d[4]} - {d[2]}: {d[3]}")
        y -= 15

    y -= 15
    p.setFont("Helvetica-Bold", 10)
    p.drawString(50, y, "Abonos:")
    y -= 20
    total_abonos = 0
    for a in abonos:
        total_abonos += a[2]
        p.drawString(60, y, f"{a[3]} - ${a[2]:.2f} - {a[4]}")
        y -= 15

    y -= 25
    p.setFont("Helvetica-Bold", 10)
    total_deuda = total_entregado - total_devoluciones - total_abonos
    p.drawString(50, y, f"TOTAL ENTREGADO: ${total_entregado:.2f}")
    y -= 20
    p.drawString(50, y, f"TOTAL DEVOLUCIONES: ${total_devoluciones:.2f}")
    y -= 20
    p.drawString(50, y, f"TOTAL ABONOS: ${total_abonos:.2f}")
    y -= 20
    p.drawString(50, y, f"TOTAL DEUDA: ${total_deuda:.2f}")

    p.showPage()
    p.save()

    buffer.seek(0)
    return buffer


def init_db():
    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()

    # Tabla de administradores / usuarios
    c.execute("""
        CREATE TABLE IF NOT EXISTS admin (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT,
            email TEXT,
            password TEXT,
            rol TEXT DEFAULT 'usuario'
        )
    """)

    # Tabla de libros
    c.execute("""
        CREATE TABLE IF NOT EXISTS libros (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT,
            precio REAL,
            imagen TEXT
        )
    """)

    # Tabla de clientes
    c.execute("""
        CREATE TABLE IF NOT EXISTS clientes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT,
            foto TEXT
        )
    """)

    # Tabla de ventas (detalle de venta)
    c.execute("""
        CREATE TABLE IF NOT EXISTS ventas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cliente_id INTEGER,
            libro TEXT,
            cantidad INTEGER,
            precio_unitario REAL,
            fecha TEXT,
            FOREIGN KEY (cliente_id) REFERENCES clientes(id)
        )
    """)

    # Tabla de abonos
    c.execute("""
        CREATE TABLE IF NOT EXISTS abonos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cliente_id INTEGER,
            monto REAL,
            fecha TEXT,
            comentario TEXT,
            FOREIGN KEY (cliente_id) REFERENCES clientes(id)
        )
    """)

    # Tabla de devoluciones
    c.execute("""
        CREATE TABLE IF NOT EXISTS devoluciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cliente_id INTEGER,
            libro TEXT,
            cantidad INTEGER,
            fecha TEXT,
            FOREIGN KEY (cliente_id) REFERENCES clientes(id)
        )
    """)

    conn.commit()
    conn.close()

@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        conn = sqlite3.connect(DATABASE)
        c = conn.cursor()
        c.execute("SELECT * FROM admin WHERE email=? AND password=?", (email, password))
        user = c.fetchone()
        conn.close()
        if user:
            session['admin'] = user[1]
            session['rol'] = user[4]  # columna 4 = rol

            if user[4] == 'admin':
                return redirect(url_for('dashboard'))
            else:
                return redirect(url_for('dashboard_usuario'))

        else:
            flash("Credenciales incorrectas", "danger")
    return render_template('login.html')

@app.route('/dashboard')
def dashboard():
    if 'admin' not in session:
        return redirect(url_for('login'))
    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    c.execute("SELECT * FROM libros")
    libros = c.fetchall()
    conn.close()
    return render_template('dashboard.html', libros=libros)

@app.route('/agregar_libro', methods=['POST'])
def agregar_libro():
    if 'admin' not in session:
        return redirect(url_for('login'))
    nombre = request.form['nombre']
    precio = float(request.form['precio'])
    imagen = request.files['imagen']
    filename = secure_filename(imagen.filename)
    imagen.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))

    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    c.execute("INSERT INTO libros (nombre, precio, imagen) VALUES (?, ?, ?)", (nombre, precio, filename))
    conn.commit()
    conn.close()
    return redirect(url_for('dashboard'))

@app.route('/libros')
def vista_libros():
    if 'admin' not in session:
        return redirect(url_for('login'))
    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    c.execute("SELECT * FROM libros")
    libros = c.fetchall()
    conn.close()
    return render_template('libros.html', libros=libros)

@app.route('/usuarios')
def vista_usuarios():
    if 'admin' not in session:
        return redirect(url_for('login'))
    return render_template('usuarios.html')

@app.route('/registrar_usuario', methods=['POST'])
def registrar_usuario():
    if 'admin' not in session:
        return redirect(url_for('login'))

    nombre = request.form['nombre']
    email = request.form['email']
    password = request.form['password']

    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    c.execute("INSERT INTO admin (nombre, email, password) VALUES (?, ?, ?)", (nombre, email, password))
    conn.commit()
    conn.close()

    return redirect(url_for('dashboard'))

@app.route('/dashboard_usuario')
def dashboard_usuario():
    if 'admin' not in session or session.get('rol') != 'usuario':
        return redirect(url_for('login'))

    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    c.execute("SELECT * FROM libros")
    libros = c.fetchall()
    conn.close()
    return render_template('dashboard_usuario.html', libros=libros)

@app.route('/clientes')
def clientes():
    if 'admin' not in session or session.get('rol') != 'usuario':
        return redirect(url_for('login'))

    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    c.execute("SELECT * FROM clientes")
    clientes = c.fetchall()
    conn.close()
    return render_template('clientes.html', clientes=clientes)

@app.route('/agregar_cliente', methods=['POST'])
def agregar_cliente():
    if 'admin' not in session or session.get('rol') != 'usuario':
        return redirect(url_for('login'))

    nombre = request.form['nombre']
    foto_file = request.files.get('foto')
    foto_nombre = None

    if foto_file and foto_file.filename != '':
        from werkzeug.utils import secure_filename
        foto_nombre = secure_filename(foto_file.filename)
        foto_file.save(os.path.join('static/images', foto_nombre))

    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    c.execute("INSERT INTO clientes (nombre, foto) VALUES (?, ?)", (nombre, foto_nombre))
    conn.commit()
    conn.close()
    return redirect(url_for('clientes'))


@app.route('/ver_cliente/<int:cliente_id>')
def ver_cliente(cliente_id):
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM clientes WHERE id = ?", (cliente_id,))
    cliente = cursor.fetchone()

    cursor.execute("SELECT cantidad, precio_unitario, descuento FROM ventas WHERE cliente_id = ?", (cliente_id,))
    ventas = cursor.fetchall()

    total_entregado = sum(
        cantidad * (precio_unitario * (1 - descuento / 100))
        for cantidad, precio_unitario, descuento in ventas
    )

    cursor.execute("SELECT libro, cantidad FROM devoluciones WHERE cliente_id = ?", (cliente_id,))
    devoluciones = cursor.fetchall()
    total_devoluciones = sum(cant * precio_del_libro(lib) for lib, cant in devoluciones)

    cursor.execute("SELECT monto FROM abonos WHERE cliente_id = ?", (cliente_id,))
    abonos = cursor.fetchall()
    total_abonos = sum(a[0] for a in abonos)

    conn.close()

    total_deuda = total_entregado - total_devoluciones - total_abonos

    return render_template(
        'ver_cliente.html',
        cliente=cliente,
        total_entregado=round(total_entregado, 2),
        total_devoluciones=round(total_devoluciones, 2),
        total_abonos=round(total_abonos, 2),
        total_deuda=round(total_deuda, 2)
    )


@app.route('/eliminar_cliente/<int:cliente_id>', methods=['POST'])
def eliminar_cliente(cliente_id):
    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    c.execute("DELETE FROM clientes WHERE id = ?", (cliente_id,))
    conn.commit()
    conn.close()
    flash("Cliente eliminado", "success")
    return redirect(url_for('clientes'))

@app.route('/cliente/<int:cliente_id>/agregar_entrega', methods=['POST'])
def agregar_entrega(cliente_id):
    fecha = request.form['fecha']
    nota = request.form['nota']
    materias = request.form.getlist('materias[]')
    cantidades = request.form.getlist('cantidades[]')
    descuentos = request.form.getlist('descuentos[]')
    precios = request.form.getlist('precios[]')
    imagen = request.files.get('imagen')

    guardar_entrega(cliente_id, fecha, nota, materias, cantidades, descuentos, precios, imagen)

    return redirect(url_for('ver_entregas', cliente_id=cliente_id))


@app.route('/cliente/<int:cliente_id>/agregar_devolucion', methods=['POST'])
def agregar_devolucion(cliente_id):
    fecha = request.form['fecha']
    materias = request.form.getlist('materias[]')
    cantidades = request.form.getlist('cantidades[]')

    for materia, cantidad in zip(materias, cantidades):
        if materia and int(cantidad) > 0:
            guardar_devolucion(cliente_id, fecha, materia, int(cantidad))

    return redirect(url_for('ver_cliente', cliente_id=cliente_id))

@app.route('/cliente/<int:cliente_id>/agregar_abono', methods=['POST'])
def agregar_abono(cliente_id):
    fecha = request.form['fecha']
    valor = float(request.form['valor'])
    referencia = request.form['referencia']
    
    archivo = request.files.get('comprobante')
    nombre_archivo = None

    if archivo and archivo.filename:
        from werkzeug.utils import secure_filename
        nombre_archivo = secure_filename(archivo.filename)
        archivo.save(os.path.join('static/comprobantes', nombre_archivo))

    # Guarda el abono incluyendo el nombre del archivo
    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    c.execute("""
        INSERT INTO abonos (cliente_id, monto, fecha, comentario, comprobante)
        VALUES (?, ?, ?, ?, ?)
    """, (cliente_id, valor, fecha, referencia, nombre_archivo))
    conn.commit()
    conn.close()

    return redirect(url_for('ver_cliente', cliente_id=cliente_id))


@app.route('/cliente/<int:cliente_id>/descargar_pdf')
def descargar_pdf(cliente_id):
    cliente = obtener_cliente(cliente_id)
    entregas = obtener_entregas(cliente_id)
    devoluciones = obtener_devoluciones(cliente_id)
    abonos = obtener_abonos(cliente_id)
    pdf = generar_estado_cuenta_pdf(cliente, entregas, devoluciones, abonos)
    return send_file(pdf, download_name=f'estado_cuenta_{cliente[1]}.pdf', as_attachment=True)

@app.route('/cliente/<int:cliente_id>/eliminar_nota', methods=['POST'])
def eliminar_nota(cliente_id):
    nota = request.form['nota']
    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    c.execute("DELETE FROM ventas WHERE cliente_id = ? AND nota = ?", (cliente_id, nota))
    conn.commit()
    conn.close()
    flash("Nota eliminada correctamente", "success")
    return redirect(url_for('ver_entregas', cliente_id=cliente_id))

@app.route('/cliente/<int:cliente_id>/eliminar_nota_devolucion', methods=['POST'])
def eliminar_nota_devolucion(cliente_id):
    nota = request.form['nota']

    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    c.execute("DELETE FROM devoluciones WHERE cliente_id = ? AND fecha = ?", (cliente_id, nota))
    conn.commit()
    conn.close()

    flash("Devolución eliminada correctamente", "success")
    return redirect(url_for('ver_devoluciones', cliente_id=cliente_id))

@app.route('/cliente/<int:cliente_id>/eliminar_comprobante/<int:abono_id>', methods=['POST'])
def eliminar_comprobante(cliente_id, abono_id):
    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()

    # Obtener nombre del archivo
    c.execute("SELECT comprobante FROM abonos WHERE id = ?", (abono_id,))
    resultado = c.fetchone()
    if resultado and resultado[0]:
        archivo_path = os.path.join('static/comprobantes', resultado[0])
        if os.path.exists(archivo_path):
            os.remove(archivo_path)

    # Borrar referencia al archivo en la base de datos
    c.execute("UPDATE abonos SET comprobante = NULL WHERE id = ?", (abono_id,))
    conn.commit()
    conn.close()

    flash("Comprobante eliminado correctamente", "success")
    return redirect(url_for('ver_abonos', cliente_id=cliente_id))

@app.route('/cliente/<int:cliente_id>/eliminar_abono/<int:abono_id>', methods=['POST'])
def eliminar_abono(cliente_id, abono_id):
    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()

    # Obtener comprobante para eliminar archivo
    c.execute("SELECT comprobante FROM abonos WHERE id = ?", (abono_id,))
    resultado = c.fetchone()
    if resultado and resultado[0]:
        archivo_path = os.path.join('static/comprobantes', resultado[0])
        if os.path.exists(archivo_path):
            os.remove(archivo_path)

    # Eliminar el abono completo
    c.execute("DELETE FROM abonos WHERE id = ?", (abono_id,))
    conn.commit()
    conn.close()

    flash("Abono eliminado correctamente", "success")
    return redirect(url_for('ver_abonos', cliente_id=cliente_id))


@app.route('/logout')
def logout():
    session.pop('admin', None)
    return redirect(url_for('login'))

@app.route('/cliente/<int:cliente_id>/entregas', methods=['GET', 'POST'])
def ver_entregas(cliente_id):
    conn = sqlite3.connect('database/editorial.db')
    cursor = conn.cursor()

    # Obtener datos del cliente
    cursor.execute("SELECT * FROM clientes WHERE id = ?", (cliente_id,))
    cliente = cursor.fetchone()

    # Obtener libros del sistema
    cursor.execute("SELECT nombre, precio FROM libros")
    libros_db = cursor.fetchall()
    libros = [(i, libro[0], libro[1]) for i, libro in enumerate(libros_db)]  # (índice, nombre, precio)

    # Obtener entregas
    cursor.execute("""
        SELECT id, cliente_id, libro, cantidad, precio_unitario, fecha, nota, descuento, imagen
        FROM ventas
        WHERE cliente_id = ?
        ORDER BY fecha DESC
    """, (cliente_id,))
    entregas_raw = cursor.fetchall()

    # Agrupar por nota
    notas_dict = {}
    for e in entregas_raw:
        nota = e[6]
        if nota not in notas_dict:
            notas_dict[nota] = {
                'fecha': e[5],
                'libros': {},
                'descuento': {},
                'imagen': e[8]
            }
        notas_dict[nota]['libros'][e[2]] = e[3]  # cantidades
        notas_dict[nota]['descuento'][e[2]] = e[7]  # descuentos por libro

    # Generar estructura organizada para la vista
    entregas = []
    for nota, datos in notas_dict.items():
        fila = {
            'nota': nota,
            'fecha': datos['fecha'],
            'cantidades': [],
            'total': 0,
            'descuentos': [],
            'imagen': datos['imagen']
        }
        for _, nombre_libro, precio_libro in libros:
            cantidad = datos['libros'].get(nombre_libro, 0)
            descuento = datos['descuento'].get(nombre_libro, 0)
            precio_final = precio_libro * (1 - descuento / 100)
            subtotal = cantidad * precio_final
            fila['cantidades'].append(cantidad if cantidad > 0 else "-")
            fila['descuentos'].append(f"{descuento}%" if cantidad > 0 else "-")
            fila['total'] += subtotal
        entregas.append(fila)

    conn.close()
    return render_template('ver_entregas.html', cliente=cliente, libros=libros, entregas=entregas)


@app.route('/cliente/<int:cliente_id>/devoluciones')
def ver_devoluciones(cliente_id):
    cliente = obtener_cliente(cliente_id)
    devoluciones = obtener_devoluciones(cliente_id)

    # Obtener lista de libros
    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    c.execute("SELECT * FROM libros")
    libros = c.fetchall()
    conn.close()

    # Total general
    total_general = 0.0

    # Calcular totales por libro
    totales = {}
    for libro in libros:
        nombre = libro[1]
        precio = libro[2]
        cantidad_total = sum(d[3] for d in devoluciones if d[2] == nombre)
        totales[nombre] = {'cantidad': cantidad_total}
        total_general += cantidad_total * precio

    # Agrupar devoluciones por fecha
    detalle_notas = {}
    for d in devoluciones:
        fecha = d[4]  # fecha
        libro = d[2]  # nombre del libro
        cantidad = d[3]

        if fecha not in detalle_notas:
            detalle_notas[fecha] = {}

        if libro in detalle_notas[fecha]:
            detalle_notas[fecha][libro] += cantidad
        else:
            detalle_notas[fecha][libro] = cantidad

    return render_template(
        'ver_devoluciones.html',
        cliente=cliente,
        devoluciones=devoluciones,
        libros=libros,
        total_general=total_general,
        detalle_notas=detalle_notas,
        totales=totales
    )


@app.route('/cliente/<int:cliente_id>/abonos')
def ver_abonos(cliente_id):
    cliente = obtener_cliente(cliente_id)
    abonos = obtener_abonos(cliente_id)
    return render_template('ver_abonos.html', cliente=cliente, abonos=abonos)

def precio_del_libro(nombre_libro):
    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    c.execute("SELECT precio FROM libros WHERE nombre = ?", (nombre_libro,))
    resultado = c.fetchone()
    conn.close()
    return resultado[0] if resultado else 0.0


@app.route('/proveedores')
def vista_proveedores():
    conn = sqlite3.connect(DATABASE)  # ✅ Uso correcto de la base de datos principal
    c = conn.cursor()
    c.execute("SELECT * FROM proveedores")
    proveedores = c.fetchall()
    conn.close()
    return render_template('proveedores.html', proveedores=proveedores)


@app.route('/agregar_proveedor', methods=['POST'])
def agregar_proveedor():
    if 'admin' not in session:
        return redirect(url_for('login'))

    nombre = request.form['nombre']
    contacto = request.form['contacto']
    telefono = request.form['telefono']
    direccion = request.form['direccion']

    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    c.execute("INSERT INTO proveedores (nombre, contacto, telefono, direccion) VALUES (?, ?, ?, ?)",
              (nombre, contacto, telefono, direccion))
    conn.commit()
    conn.close()

    flash("Proveedor agregado correctamente", "success")
    return redirect(url_for('vista_proveedores'))

@app.route('/proveedor/<int:proveedor_id>')
def detalle_proveedor(proveedor_id):
    if 'admin' not in session:
        return redirect(url_for('login'))

    conn = sqlite3.connect('database/editorial.db')
    c = conn.cursor()

    # Obtener datos del proveedor
    c.execute("SELECT * FROM proveedores WHERE id = ?", (proveedor_id,))
    proveedor = c.fetchone()

    # Obtener deudas del proveedor
    c.execute("SELECT fecha, valor, descripcion FROM deudas_proveedor WHERE proveedor_id = ?", (proveedor_id,))
    deudas = c.fetchall()

    # Obtener abonos realizados
    c.execute("SELECT fecha, valor, descripcion FROM abonos_proveedor WHERE proveedor_id = ?", (proveedor_id,))
    abonos = c.fetchall()

    # Calcular resumen
    total_deuda = sum([d[1] for d in deudas])
    total_abonos = sum([a[1] for a in abonos])
    saldo_pendiente = total_deuda - total_abonos

    conn.close()

    return render_template("detalle_proveedor.html", proveedor=proveedor, deudas=deudas, abonos=abonos,
                           total_deuda=total_deuda, total_abonos=total_abonos, saldo_pendiente=saldo_pendiente)

@app.route('/proveedor/<int:proveedor_id>/agregar_deuda', methods=['POST'])
def agregar_deuda_proveedor(proveedor_id):
    fecha = request.form['fecha']
    valor = float(request.form['valor'])
    descripcion = request.form['descripcion']

    conn = sqlite3.connect('database/editorial.db')
    c = conn.cursor()
    c.execute("INSERT INTO deudas_proveedor (proveedor_id, fecha, valor, descripcion) VALUES (?, ?, ?, ?)",
              (proveedor_id, fecha, valor, descripcion))
    conn.commit()
    conn.close()

    flash("Deuda agregada correctamente", "success")
    return redirect(url_for('detalle_proveedor', proveedor_id=proveedor_id))

@app.route('/proveedor/<int:proveedor_id>/agregar_abono', methods=['POST'])
def agregar_abono_proveedor(proveedor_id):
    fecha = request.form['fecha']
    valor = float(request.form['valor'])
    descripcion = request.form['descripcion']

    conn = sqlite3.connect('database/editorial.db')
    c = conn.cursor()
    c.execute("INSERT INTO abonos_proveedor (proveedor_id, fecha, valor, descripcion) VALUES (?, ?, ?, ?)",
              (proveedor_id, fecha, valor, descripcion))
    conn.commit()
    conn.close()

    flash("Abono registrado correctamente", "success")
    return redirect(url_for('detalle_proveedor', proveedor_id=proveedor_id))

@app.route('/eliminar_proveedor/<int:proveedor_id>', methods=['POST'])
def eliminar_proveedor(proveedor_id):
    conn = sqlite3.connect('database/editorial.db')
    c = conn.cursor()
    c.execute("DELETE FROM proveedores WHERE id=?", (proveedor_id,))
    conn.commit()
    conn.close()
    flash("Proveedor eliminado exitosamente", "success")
    return redirect(url_for('vista_proveedores'))


@app.route('/cliente/<int:cliente_id>/descargar_excel')
def descargar_excel(cliente_id):
    conn = sqlite3.connect('database/editorial.db')
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM clientes WHERE id = ?", (cliente_id,))
    cliente = cursor.fetchone()

    cursor.execute("""
        SELECT nota, fecha, libro, cantidad, precio_unitario, descuento
        FROM ventas
        WHERE cliente_id = ?
        ORDER BY fecha ASC
    """, (cliente_id,))
    ventas = cursor.fetchall()

    cursor.execute("""
        SELECT fecha, monto
        FROM abonos
        WHERE cliente_id = ?
        ORDER BY fecha ASC
    """, (cliente_id,))
    abonos = cursor.fetchall()

    cursor.execute("""
        SELECT fecha, libro, cantidad
        FROM devoluciones
        WHERE cliente_id = ?
        ORDER BY fecha ASC
    """, (cliente_id,))
    devoluciones = [(None, *row) for row in cursor.fetchall()]

    materias_presentes = sorted({v[2] for v in ventas + devoluciones})
    wb = openpyxl.Workbook()
    hoja = wb.active
    hoja.title = "Estado de Cuenta"

    # Estilos
    borde = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))

    def aplicar_estilo_fila(fila_celdas, color_hex, negrita=True):
        for celda in fila_celdas:
            celda.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
            celda.font = Font(bold=negrita)
            celda.border = borde

    # ENCABEZADO
    encabezado = ['Nota', 'Fecha'] + materias_presentes + ['TOTAL LIBROS', 'TOTAL USD']
    hoja.append(encabezado)

    # Calcular precios con descuento
    precios_desc = []
    for m in materias_presentes:
        for v in ventas:
            if v[2] == m:
                precios_desc.append(round(v[4] * (1 - v[5] / 100), 2))
                break
        else:
            precios_desc.append(0.0)

    # Agrupar ventas
    agrupado = {}
    for nota, fecha, libro, cantidad, precio, descuento in ventas:
        if nota not in agrupado:
            agrupado[nota] = {'fecha': fecha, 'libros': {}}
        agrupado[nota]['libros'][libro] = agrupado[nota]['libros'].get(libro, 0) + cantidad

    # Agregar filas de ventas
    for nota, datos in agrupado.items():
        fila = [nota, datos['fecha']]
        for m in materias_presentes:
            cant = datos['libros'].get(m, 0)
            fila.append(cant if cant > 0 else '')
        fila += ['', '']  # Total Libros y USD vacíos para evitar sobrecarga visual
        hoja.append(fila)

    # DEVOLUCIONES
    hoja.append(['DEVOLUCIONES', '', *['' for _ in materias_presentes], '', ''])
    fila_dev = [''] * len(encabezado)
    for nota, fecha, libro, cantidad in devoluciones:
        if libro in materias_presentes:
            idx = materias_presentes.index(libro) + 2
            fila_dev[idx] = -cantidad if cantidad > 0 else ''
    hoja.append(fila_dev)
    aplicar_estilo_fila(hoja[hoja.max_row], "FFCCCC")

    # TOTAL LIBROS Y USD
    fila_total = ['TOTAL', '']
    for m in materias_presentes:
        entregado = sum(datos['libros'].get(m, 0) for datos in agrupado.values())
        devuelto = sum(d[3] for d in devoluciones if d[2] == m)
        total = entregado - devuelto
        fila_total.append(total)
    total_libros = sum(fila_total[2:2+len(materias_presentes)])
    total_usd = round(sum(fila_total[2+i] * precios_desc[i] for i in range(len(materias_presentes))), 2)
    fila_total += [total_libros, total_usd]
    hoja.append(fila_total)
    aplicar_estilo_fila(hoja[hoja.max_row], "FCD5B4")

    # SUB USD
    fila_sub = ['SUB USD', '']
    for i, m in enumerate(materias_presentes):
        cantidad = fila_total[2+i]
        fila_sub.append(round(cantidad * precios_desc[i], 2))
    fila_sub += ['', total_usd]
    hoja.append(fila_sub)
    aplicar_estilo_fila(hoja[hoja.max_row], "B4C6E7")

    # ABONOS
    hoja.append([''] * len(encabezado))
    hoja.append(['ABONOS', 'Fecha', 'Monto'] + [''] * (len(encabezado)-3))
    aplicar_estilo_fila(hoja[hoja.max_row], "D9EAD3")
    total_abonos = 0
    for fecha, monto in abonos:
        hoja.append(['', fecha, monto] + [''] * (len(encabezado)-3))
        total_abonos += monto
    hoja.append(['', 'TOTAL ABONOS', total_abonos] + [''] * (len(encabezado)-3))
    aplicar_estilo_fila(hoja[hoja.max_row], "FFE599")

    hoja.append(['', 'TOTAL USD', total_usd] + [''] * (len(encabezado)-3))
    aplicar_estilo_fila(hoja[hoja.max_row], "D9D2E9")

    hoja.append(['', 'SALDO', round(total_usd - total_abonos, 2)] + [''] * (len(encabezado)-3))
    aplicar_estilo_fila(hoja[hoja.max_row], "FFF2CC")

    # AJUSTES FINALES
    for col in hoja.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        hoja.column_dimensions[col[0].column_letter].width = max_length + 2

    for row in hoja.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = borde

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    nombre_archivo = f"estado_cuenta_{cliente[1].replace(' ', '_')}.xlsx"
    return send_file(output, as_attachment=True, download_name=nombre_archivo,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


#end



if __name__ == '__main__':
    os.makedirs('database', exist_ok=True)
    os.makedirs('static/images', exist_ok=True)
    os.makedirs('templates', exist_ok=True)
    init_db()

    # Crear usuario administrador por defecto
    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    c.execute("SELECT * FROM admin WHERE email = 'admin@editorial.com'")
    if not c.fetchone():
        c.execute("INSERT INTO admin (nombre, email, password, rol) VALUES (?, ?, ?, ?)", 
          ('Admin', 'admin@editorial.com', 'Vero69', 'admin'))

        conn.commit()
    conn.close()

    app.run(debug=True)
